from __future__ import annotations

import hashlib
import io
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Iterable

from openpyxl import load_workbook

NON_WORKING_ALIASES = {"VAC": "VA", "INC": "IN", "C": "COMP"}

def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value).strip())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip().upper()

def normalize_key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", normalize_text(value))

def canonical_plan_name(value: Any) -> str:
    return re.sub(r"^\d+\s*[-–—:._/|]+\s*", "", normalize_text(value)).strip()

def normalize_order(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        text = re.sub(r"\.0+$", "", text)
    return normalize_text(text)

def normalize_turn(value: Any) -> str:
    code = normalize_text(value)
    return NON_WORKING_ALIASES.get(code, code)

def as_date(value: Any) -> date | None:
    if value in (None, ""): return None
    if isinstance(value, datetime): return value.date()
    if isinstance(value, date): return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try: return datetime.strptime(text, fmt).date()
        except ValueError: pass
    return None

def as_datetime(value: Any) -> datetime | None:
    if value in (None, ""): return None
    if isinstance(value, datetime): return value
    if isinstance(value, date): return datetime.combine(value, datetime.min.time())
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d", "%d/%m/%Y"):
        try: return datetime.strptime(text, fmt)
        except ValueError: pass
    return None

def as_float(value: Any, default: float = 0.0) -> float:
    if value in (None, ""): return default
    if isinstance(value, (int, float)): return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".") if "," in str(value) else str(value).strip()
    try: return float(text)
    except ValueError: return default

def sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()

def stable_sha256(parts: Iterable[Any]) -> str:
    payload = "|".join(normalize_text(p) for p in parts)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()

@dataclass(slots=True)
class ParsedWorkbook:
    rows: list[dict[str, Any]] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

def workbook_from_bytes(content: bytes):
    return load_workbook(io.BytesIO(content), read_only=True, data_only=True)

def find_header_row(ws, required_columns: Iterable[str], max_rows: int = 50) -> tuple[int, dict[str, int]]:
    required = {normalize_key(c) for c in required_columns}
    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        mapping = {normalize_key(v): i for i, v in enumerate(row, start=1) if normalize_key(v)}
        if required.issubset(mapping.keys()):
            return row_index, mapping
    raise ValueError(f"No se encontraron columnas requeridas: {', '.join(required_columns)}")

def header_mapping(ws, header_row: int) -> dict[str, int]:
    row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    return {normalize_key(v): i for i, v in enumerate(row, start=1) if normalize_key(v)}

def cell_by_header(row: tuple[Any, ...], mapping: dict[str, int], *aliases: str) -> Any:
    for alias in aliases:
        idx = mapping.get(normalize_key(alias))
        if idx and idx <= len(row):
            return row[idx - 1]
    return None
