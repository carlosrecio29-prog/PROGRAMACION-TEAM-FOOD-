from __future__ import annotations

import io
import json
import os
from datetime import datetime
from typing import Any
from urllib.parse import quote

import requests
from google.auth.transport.requests import Request as GoogleAuthRequest
from google.oauth2 import service_account
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from backend.parsers.common import header_mapping, cell_by_header, normalize_text

XLSX_MIME="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DEFAULT_TEAM_FOOD_SHEET_ID="1pgUoKEX7FAVOwB4c9aUPdka_hsOr7ULZ1ntUqIOcNk0"

class GoogleSheetSyncError(RuntimeError):
    pass

def _config()->dict[str,Any]:
    return {
        "sheet_id":(os.getenv("TEAM_FOOD_SHEET_ID") or DEFAULT_TEAM_FOOD_SHEET_ID).strip(),
        "raw_credentials":(os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON") or "").strip(),
    }

def _credentials():
    cfg=_config()
    if not cfg["raw_credentials"]:
        raise GoogleSheetSyncError("GOOGLE_SERVICE_ACCOUNT_JSON no está configurado")
    try:
        info=json.loads(cfg["raw_credentials"])
        if isinstance(info.get("private_key"),str):
            info["private_key"]=info["private_key"].replace("\\n","\n")
        creds=service_account.Credentials.from_service_account_info(
            info,
            scopes=[
                "https://www.googleapis.com/auth/drive.readonly",
                "https://www.googleapis.com/auth/spreadsheets",
            ],
        )
        creds.refresh(GoogleAuthRequest())
        return creds
    except Exception as exc:
        raise GoogleSheetSyncError("No se pudieron cargar las credenciales de Google") from exc

def _download_xlsx(creds)->bytes:
    cfg=_config()
    url=(
        f"https://www.googleapis.com/drive/v3/files/{quote(cfg['sheet_id'],safe='')}/export"
        f"?mimeType={quote(XLSX_MIME,safe='')}"
    )
    response=requests.get(url,headers={"Authorization":f"Bearer {creds.token}"},timeout=35)
    if not response.ok:
        raise GoogleSheetSyncError(f"No se pudo leer TEAM FOOD desde Google (HTTP {response.status_code})")
    return response.content

def _sheet(wb,expected:str):
    key=normalize_text(expected)
    for ws in wb.worksheets:
        if normalize_text(ws.title)==key:
            return ws
    return None

def _find_plan_targets(
    content:bytes,
    *,
    group_code:str,
    specialty:str,
    plan_name:str,
)->tuple[list[int],dict[str,int]]:
    wb=load_workbook(io.BytesIO(content),read_only=True,data_only=True)
    ws=_sheet(wb,"PLAN DE TRABAJO")
    if ws is None:
        raise GoogleSheetSyncError("No se encontró la hoja PLAN DE TRABAJO")

    header_values=next(ws.iter_rows(min_row=1,max_row=1,values_only=True))
    columns={
        normalize_text(value):idx
        for idx,value in enumerate(header_values,start=1)
        if value not in (None,"")
    }
    people_col=columns.get(normalize_text("NumeroPersonas"))
    stopped_col=columns.get(normalize_text("EquipoDetenido"))
    if people_col is None:
        raise GoogleSheetSyncError("No existe la columna NumeroPersonas en PLAN DE TRABAJO")
    if stopped_col is None:
        raise GoogleSheetSyncError("No existe la columna EquipoDetenido en PLAN DE TRABAJO")

    m=header_mapping(ws,1)
    rows=[]
    for n,row in enumerate(ws.iter_rows(min_row=2,values_only=True),start=2):
        group=str(cell_by_header(row,m,"Grupo") or "").strip()
        spec=normalize_text(cell_by_header(row,m,"Especialidad"))
        plan=normalize_text(cell_by_header(row,m,"PlanTrabajo","Plan de Trabajo"))
        if group==str(group_code).strip() and spec==normalize_text(specialty) and plan==normalize_text(plan_name):
            rows.append(n)

    return rows,{"people":people_col,"stopped":stopped_col}


def sync_plan_definition_to_google_sheet(
    *,
    group_code:str,
    specialty:str,
    plan_name:str,
    people:float|None,
    condition:str|None,
)->dict[str,Any]:
    creds=_credentials()
    cfg=_config()
    content=_download_xlsx(creds)
    plan_rows,columns=_find_plan_targets(
        content,
        group_code=group_code,
        specialty=specialty,
        plan_name=plan_name,
    )

    if not plan_rows:
        raise GoogleSheetSyncError("No se encontró el plan exacto en PLAN DE TRABAJO")

    equipment_stopped=None
    if condition=="EQUIPO DETENIDO":
        equipment_stopped="SI"
    elif condition=="OPERANDO":
        equipment_stopped="NO"

    data=[]
    people_letter=get_column_letter(columns["people"])
    stopped_letter=get_column_letter(columns["stopped"])

    for row in plan_rows:
        if people is not None:
            data.append({
                "range":f"'PLAN DE TRABAJO '!{people_letter}{row}",
                "values":[[float(people)]],
            })
        if equipment_stopped is not None:
            data.append({
                "range":f"'PLAN DE TRABAJO '!{stopped_letter}{row}",
                "values":[[equipment_stopped]],
            })

    if not data:
        return {
            "ok":True,
            "plan_rows_updated":0,
            "sheet_id":cfg["sheet_id"],
        }

    endpoint=f"https://sheets.googleapis.com/v4/spreadsheets/{quote(cfg['sheet_id'],safe='')}/values:batchUpdate"
    response=requests.post(
        endpoint,
        headers={"Authorization":f"Bearer {creds.token}","Content-Type":"application/json"},
        json={"valueInputOption":"USER_ENTERED","data":data},
        timeout=35,
    )
    if not response.ok:
        raise GoogleSheetSyncError(f"No se pudo actualizar TEAM FOOD (HTTP {response.status_code})")

    return {
        "ok":True,
        "plan_rows_updated":len(plan_rows),
        "sheet_id":cfg["sheet_id"],
        "columns":{
            "NumeroPersonas":people_letter,
            "EquipoDetenido":stopped_letter,
        },
    }
