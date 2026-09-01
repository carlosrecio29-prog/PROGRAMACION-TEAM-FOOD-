from __future__ import annotations
from datetime import date
from io import BytesIO
from html import escape

from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from backend.database import get_engine
from backend.services.query_service import get_capacity

class ProgrammingError(ValueError): pass

def _specialty_id(conn,code):
    v=conn.execute(text("SELECT id FROM mantenimiento.especialidad WHERE codigo=:c"),{"c":code.upper()}).scalar_one_or_none()
    if not v: raise ProgrammingError(f"Especialidad desconocida: {code}")
    return int(v)

def save_programming(*,date_from:date,date_to:date,specialty:str,pmp_ids:list[int],created_by=None,reason=None):
    if not pmp_ids: raise ProgrammingError("Debes seleccionar al menos un PMP")
    if date_to<date_from or (date_to-date_from).days>6: raise ProgrammingError("El rango semanal debe ser de 1 a 7 días")
    cap=get_capacity(date_from,date_to).get(specialty.upper(),{"available":0,"target":0,"standby":0})
    with get_engine().begin() as conn:
        sid=_specialty_id(conn,specialty)
        period=conn.execute(text("""INSERT INTO mantenimiento.periodo_semanal(fecha_desde,fecha_hasta) VALUES(:d,:h)
        ON CONFLICT(fecha_desde,fecha_hasta) DO UPDATE SET fecha_desde=EXCLUDED.fecha_desde RETURNING id"""),{"d":date_from,"h":date_to}).scalar_one()
        pid=conn.execute(text("""INSERT INTO mantenimiento.programacion_semanal(periodo_semanal_id,especialidad_id,estado,hh_disponibles,hh_objetivo,hh_standby,creado_por)
        VALUES(:p,:s,'GUARDADA',:a,:o,:st,:u) ON CONFLICT(periodo_semanal_id,especialidad_id) DO UPDATE SET estado='GUARDADA',
        hh_disponibles=EXCLUDED.hh_disponibles,hh_objetivo=EXCLUDED.hh_objetivo,hh_standby=EXCLUDED.hh_standby,guardado_en=NOW() RETURNING id"""),
        {"p":period,"s":sid,"a":cap["available"],"o":cap["target"],"st":cap["standby"],"u":created_by}).scalar_one()
        rows=conn.execute(text("""SELECT * FROM mantenimiento.vw_pmp_calculado WHERE pmp_id=ANY(:ids) AND especialidad=:s AND estado_orden<>'FINALIZADA'"""),{"ids":pmp_ids,"s":specialty.upper()}).mappings().all()
        if len(rows)!=len(set(pmp_ids)): raise ProgrammingError("Uno o más PMP no existen o ya están FINALIZADOS")
        incomplete=[r for r in rows if not r["datos_completos"]]
        if incomplete:
            faltantes=", ".join(sorted({x for r in incomplete for x in (r["datos_faltantes"] or [])}))
            raise ProgrammingError(f"Hay {len(incomplete)} PMP con datos incompletos ({faltantes}). Complétalos antes de programar.")
        total=sum(float(r["hh_pmp"]) for r in rows)
        if total>float(cap["target"] or 0)+.0001: raise ProgrammingError(f"La programación ({total:.1f} HH) supera la meta ({float(cap['target']):.1f} HH)")
        mx=conn.execute(text("SELECT COALESCE(MAX(numero_version),0) FROM mantenimiento.programacion_version WHERE programacion_semanal_id=:p"),{"p":pid}).scalar_one()
        ver=int(mx)+1
        conn.execute(text("UPDATE mantenimiento.programacion_version SET es_actual=FALSE WHERE programacion_semanal_id=:p AND es_actual=TRUE"),{"p":pid})
        vid=conn.execute(text("""INSERT INTO mantenimiento.programacion_version(programacion_semanal_id,numero_version,tipo,motivo,es_actual)
        VALUES(:p,:v,:t,:m,TRUE) RETURNING id"""),{"p":pid,"v":ver,"t":"INICIAL" if ver==1 else "REPROGRAMACION","m":reason}).scalar_one()
        for r in rows:
            backlog=conn.execute(text("SELECT 1 FROM mantenimiento.vw_backlog WHERE pmp_id=:p"),{"p":r["pmp_id"]}).scalar_one_or_none()
            conn.execute(text("""INSERT INTO mantenimiento.programacion_item(programacion_version_id,pmp_id,orden_id,origen,tiempo_planeado_min,personas_usar,hh_programadas,condicion_snapshot,criticidad_snapshot)
            VALUES(:v,:p,:o,:ori,:m,:per,:hh,:c,:cr)"""),{"v":vid,"p":r["pmp_id"],"o":r["orden_id"],"ori":"BACKLOG" if backlog else "MES","m":r["tiempo_planeado_min"],"per":r["personas_usar"],"hh":r["hh_pmp"],"c":r["condicion"],"cr":r["criticidad"]})
    return {"programming_id":int(pid),"version_id":int(vid),"version":ver,"hh_programmed":total,"hh_target":cap["target"]}

def programming_history(limit=50):
    with get_engine().connect() as conn:
        rows=conn.execute(text("""SELECT ps.id programming_id,pv.id version_id,pv.numero_version,pv.tipo,pv.es_actual,pv.creado_en,per.fecha_desde,per.fecha_hasta,e.codigo especialidad,
        ps.hh_disponibles,ps.hh_objetivo,ps.hh_standby,COALESCE(SUM(pi.hh_programadas),0)::float hh_programadas,COUNT(pi.id) items
        FROM mantenimiento.programacion_semanal ps JOIN mantenimiento.periodo_semanal per ON per.id=ps.periodo_semanal_id
        JOIN mantenimiento.especialidad e ON e.id=ps.especialidad_id JOIN mantenimiento.programacion_version pv ON pv.programacion_semanal_id=ps.id
        LEFT JOIN mantenimiento.programacion_item pi ON pi.programacion_version_id=pv.id GROUP BY ps.id,pv.id,per.id,e.id ORDER BY pv.creado_en DESC LIMIT :l"""),{"l":limit}).mappings().all()
    return [dict(r) for r in rows]

def programming_detail(version_id:int):
    with get_engine().connect() as conn:
        rows=conn.execute(text("""SELECT pi.id program_item_id,pi.pmp_id,om.numero_orden,a.codigo activo_codigo,a.descripcion activo_descripcion,a.area_nombre,a.linea_nombre,
        pt.nombre plan_trabajo,pi.criticidad_snapshot criticidad,pi.condicion_snapshot condicion,pi.personas_usar,pi.tiempo_planeado_min,pi.hh_programadas,pi.origen,
        COALESCE(om.estado,'PENDIENTE') estado FROM mantenimiento.programacion_item pi JOIN mantenimiento.pmp p ON p.id=pi.pmp_id
        JOIN mantenimiento.activo a ON a.id=p.activo_id JOIN mantenimiento.plan_trabajo pt ON pt.id=p.plan_trabajo_id
        LEFT JOIN mantenimiento.orden_mantenimiento om ON om.id=pi.orden_id WHERE pi.programacion_version_id=:v ORDER BY a.area_nombre,pt.nombre,a.codigo"""),{"v":version_id}).mappings().all()
    return [dict(r) for r in rows]

def close_programming(*,programming_id:int,version_id:int,reasons:dict,closed_by=None):
    with get_engine().begin() as conn:
        rows=conn.execute(text("""SELECT pi.id item_id,pi.hh_programadas,COALESCE(om.estado,'PENDIENTE') estado FROM mantenimiento.programacion_item pi
        LEFT JOIN mantenimiento.orden_mantenimiento om ON om.id=pi.orden_id WHERE pi.programacion_version_id=:v"""),{"v":version_id}).mappings().all()
        if not rows: raise ProgrammingError("La versión no tiene actividades")
        total=sum(float(r["hh_programadas"] or 0) for r in rows); final=sum(float(r["hh_programadas"] or 0) for r in rows if r["estado"]=="FINALIZADA"); pending=total-final; pct=final/total*100 if total else 0
        cid=conn.execute(text("""INSERT INTO mantenimiento.cierre_semanal(programacion_semanal_id,programacion_version_id,hh_programadas,hh_finalizadas,hh_pendientes,cumplimiento_pct,cerrado_por)
        VALUES(:p,:v,:t,:f,:pe,:pct,:u) ON CONFLICT(programacion_semanal_id,programacion_version_id) DO UPDATE SET hh_programadas=EXCLUDED.hh_programadas,
        hh_finalizadas=EXCLUDED.hh_finalizadas,hh_pendientes=EXCLUDED.hh_pendientes,cumplimiento_pct=EXCLUDED.cumplimiento_pct,cerrado_por=EXCLUDED.cerrado_por,cerrado_en=NOW() RETURNING id"""),
        {"p":programming_id,"v":version_id,"t":total,"f":final,"pe":pending,"pct":pct,"u":closed_by}).scalar_one()
        conn.execute(text("DELETE FROM mantenimiento.cierre_item WHERE cierre_semanal_id=:c"),{"c":cid})
        for r in rows:
            reason_id=None; detail=related=None
            if r["estado"]!="FINALIZADA":
                supplied=reasons.get(int(r["item_id"])) or {}; code=supplied.get("reason_code")
                if not code: raise ProgrammingError(f"Falta motivo para item {r['item_id']}")
                reason=conn.execute(text("SELECT id,requiere_detalle FROM mantenimiento.motivo_no_ejecucion WHERE codigo=:c AND activo=TRUE"),{"c":code}).mappings().one_or_none()
                if not reason: raise ProgrammingError(f"Motivo desconocido: {code}")
                detail=supplied.get("detail"); related=supplied.get("related_order")
                if reason["requiere_detalle"] and not detail: raise ProgrammingError(f"El motivo {code} requiere detalle")
                reason_id=reason["id"]
            conn.execute(text("""INSERT INTO mantenimiento.cierre_item(cierre_semanal_id,programacion_item_id,estado_ejecucion,motivo_no_ejecucion_id,detalle,orden_relacionada)
            VALUES(:c,:i,:s,:r,:d,:o)"""),{"c":cid,"i":r["item_id"],"s":r["estado"],"r":reason_id,"d":detail,"o":related})
        conn.execute(text("UPDATE mantenimiento.programacion_semanal SET estado='CERRADA',cerrado_en=NOW() WHERE id=:id"),{"id":programming_id})
    return {"closure_id":int(cid),"hh_programmed":total,"hh_finalized":final,"hh_pending":pending,"compliance_pct":pct}


def _programming_export_data(version_id:int):
    with get_engine().connect() as conn:
        header=conn.execute(text("""SELECT
          ps.id programming_id,pv.id version_id,pv.numero_version,pv.tipo,pv.creado_en,
          per.fecha_desde,per.fecha_hasta,e.codigo especialidad,
          ps.hh_disponibles,ps.hh_objetivo,ps.hh_standby,
          COALESCE(SUM(pi.hh_programadas),0)::float hh_programadas,
          COUNT(pi.id)::int items
        FROM mantenimiento.programacion_semanal ps
        JOIN mantenimiento.periodo_semanal per ON per.id=ps.periodo_semanal_id
        JOIN mantenimiento.especialidad e ON e.id=ps.especialidad_id
        JOIN mantenimiento.programacion_version pv ON pv.programacion_semanal_id=ps.id
        LEFT JOIN mantenimiento.programacion_item pi ON pi.programacion_version_id=pv.id
        WHERE pv.id=:v
        GROUP BY ps.id,pv.id,per.id,e.id"""),{"v":version_id}).mappings().one_or_none()
        if not header:
            raise ProgrammingError("Versión de programación no encontrada")

        rows=conn.execute(text("""SELECT
          om.numero_orden,a.area_nombre,a.linea_nombre,a.codigo activo_codigo,
          a.descripcion activo_descripcion,p.titulo actividad,pt.nombre plan_trabajo,
          pi.criticidad_snapshot criticidad,pi.condicion_snapshot condicion,
          pi.personas_usar,pi.tiempo_planeado_min,pi.hh_programadas,pi.origen,
          COALESCE(om.estado,'PENDIENTE') estado
        FROM mantenimiento.programacion_item pi
        JOIN mantenimiento.pmp p ON p.id=pi.pmp_id
        JOIN mantenimiento.activo a ON a.id=p.activo_id
        JOIN mantenimiento.plan_trabajo pt ON pt.id=p.plan_trabajo_id
        LEFT JOIN mantenimiento.orden_mantenimiento om ON om.id=pi.orden_id
        WHERE pi.programacion_version_id=:v
        ORDER BY a.area_nombre,pt.nombre,p.titulo,a.codigo"""),{"v":version_id}).mappings().all()
    return dict(header),[dict(r) for r in rows]


def export_programming_excel(version_id:int):
    header,rows=_programming_export_data(version_id)
    wb=Workbook()
    ws=wb.active
    ws.title="Programación semanal"
    ws.sheet_view.showGridLines=False
    ws.freeze_panes="A9"
    ws.page_setup.orientation="landscape"
    ws.page_setup.paperSize=ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth=1
    ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    ws.print_title_rows="1:8"
    ws.page_margins.left=0.25
    ws.page_margins.right=0.25
    ws.page_margins.top=0.35
    ws.page_margins.bottom=0.35

    navy="17365D"; blue="2F75B5"; light="D9EAF7"; pale="F3F6FA"
    green="E2F0D9"; white="FFFFFF"; dark="1F2937"; gray="6B7280"
    thin=Side(style="thin",color="D9E2F3")

    ws.merge_cells("A1:N1")
    ws["A1"]="PROGRAMACIÓN SEMANAL DE MANTENIMIENTO"
    ws["A1"].font=Font(size=18,bold=True,color=white)
    ws["A1"].fill=PatternFill("solid",fgColor=navy)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=30

    ws.merge_cells("A2:N2")
    ws["A2"]=f"Semana {header['fecha_desde']:%d/%m/%Y} al {header['fecha_hasta']:%d/%m/%Y}  |  Especialidad: {header['especialidad']}  |  Versión: {header['numero_version']}"
    ws["A2"].font=Font(size=10,bold=True,color=dark)
    ws["A2"].alignment=Alignment(horizontal="center")

    cards=[
      ("A4:C4","A5:C6","H-H DISPONIBLES",float(header["hh_disponibles"] or 0)),
      ("D4:F4","D5:F6","META 80%",float(header["hh_objetivo"] or 0)),
      ("G4:I4","G5:I6","H-H PROGRAMADAS",float(header["hh_programadas"] or 0)),
      ("J4:L4","J5:L6","STANDBY 20%",float(header["hh_standby"] or 0)),
      ("M4:N4","M5:N6","ACTIVIDADES",int(header["items"] or 0)),
    ]
    for title_range,value_range,label,value in cards:
        ws.merge_cells(title_range);ws.merge_cells(value_range)
        t=ws[title_range.split(":")[0]];v=ws[value_range.split(":")[0]]
        t.value=label;t.font=Font(size=9,bold=True,color=gray);t.fill=PatternFill("solid",fgColor=pale);t.alignment=Alignment(horizontal="center")
        v.value=value;v.font=Font(size=16,bold=True,color=navy);v.fill=PatternFill("solid",fgColor=light);v.alignment=Alignment(horizontal="center",vertical="center")
        v.number_format='0.0'

    headers=["OT","Área","Línea","Código equipo","Descripción equipo","Actividad","Plan de trabajo","Crit.","Condición","Personas","Tiempo min","H-H","Origen","Estado"]
    for c,label in enumerate(headers,1):
        cell=ws.cell(row=8,column=c,value=label)
        cell.font=Font(bold=True,color=white,size=9)
        cell.fill=PatternFill("solid",fgColor=blue)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=Border(left=thin,right=thin,top=thin,bottom=thin)

    for r_idx,row in enumerate(rows,9):
        values=[
          row.get("numero_orden") or "",row.get("area_nombre") or "",row.get("linea_nombre") or "",
          row.get("activo_codigo") or "",row.get("activo_descripcion") or "",row.get("actividad") or "",
          row.get("plan_trabajo") or "",row.get("criticidad") or "",row.get("condicion") or "",
          float(row["personas_usar"]) if row.get("personas_usar") is not None else "",
          float(row["tiempo_planeado_min"]) if row.get("tiempo_planeado_min") is not None else "",
          float(row["hh_programadas"]) if row.get("hh_programadas") is not None else "",
          row.get("origen") or "",row.get("estado") or ""
        ]
        for c_idx,value in enumerate(values,1):
            cell=ws.cell(row=r_idx,column=c_idx,value=value)
            cell.font=Font(size=8,color=dark)
            cell.alignment=Alignment(vertical="top",wrap_text=True)
            cell.border=Border(left=thin,right=thin,top=thin,bottom=thin)
            if r_idx%2==0:cell.fill=PatternFill("solid",fgColor="F8FAFC")
        ws.cell(r_idx,10).number_format='0'
        ws.cell(r_idx,11).number_format='0.0'
        ws.cell(r_idx,12).number_format='0.0'

    widths=[16,18,17,21,34,34,34,8,20,10,11,10,10,12]
    for idx,width in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(idx)].width=width
    ws.auto_filter.ref=f"A8:N{max(8,8+len(rows))}"
    ws.print_area=f"A1:N{max(8,8+len(rows))}"

    out=BytesIO()
    wb.save(out)
    out.seek(0)
    filename=f"programacion_{header['especialidad']}_{header['fecha_desde']:%Y%m%d}_{header['fecha_hasta']:%Y%m%d}_v{header['numero_version']}.xlsx"
    return out.getvalue(),filename


def export_programming_pdf(version_id:int):
    header,rows=_programming_export_data(version_id)
    out=BytesIO()
    doc=SimpleDocTemplate(
        out,pagesize=landscape(A4),
        leftMargin=8*mm,rightMargin=8*mm,topMargin=8*mm,bottomMargin=8*mm,
        title="Programación semanal de mantenimiento"
    )
    styles=getSampleStyleSheet()
    title_style=ParagraphStyle("ReportTitle",parent=styles["Heading1"],fontName="Helvetica-Bold",
        fontSize=15,leading=18,textColor=colors.HexColor("#17365D"),alignment=TA_CENTER,spaceAfter=4)
    subtitle_style=ParagraphStyle("ReportSub",parent=styles["Normal"],fontName="Helvetica",
        fontSize=8,leading=10,textColor=colors.HexColor("#4B5563"),alignment=TA_CENTER,spaceAfter=7)
    cell_style=ParagraphStyle("Cell",parent=styles["Normal"],fontName="Helvetica",fontSize=5.7,leading=7)
    cell_center=ParagraphStyle("CellCenter",parent=cell_style,alignment=TA_CENTER)
    small_bold=ParagraphStyle("SmallBold",parent=cell_style,fontName="Helvetica-Bold")

    story=[
      Paragraph("PROGRAMACIÓN SEMANAL DE MANTENIMIENTO",title_style),
      Paragraph(
        f"Semana {header['fecha_desde']:%d/%m/%Y} al {header['fecha_hasta']:%d/%m/%Y} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"Especialidad: <b>{header['especialidad']}</b> &nbsp;&nbsp;|&nbsp;&nbsp; Versión: {header['numero_version']}",
        subtitle_style
      )
    ]

    summary=[
      [Paragraph("H-H DISPONIBLES",small_bold),Paragraph("META 80%",small_bold),
       Paragraph("H-H PROGRAMADAS",small_bold),Paragraph("STANDBY 20%",small_bold),Paragraph("ACTIVIDADES",small_bold)],
      [f"{float(header['hh_disponibles'] or 0):.1f}",f"{float(header['hh_objetivo'] or 0):.1f}",
       f"{float(header['hh_programadas'] or 0):.1f}",f"{float(header['hh_standby'] or 0):.1f}",str(int(header['items'] or 0))]
    ]
    summary_table=Table(summary,colWidths=[54*mm,54*mm,54*mm,54*mm,45*mm],rowHeights=[7*mm,8*mm])
    summary_table.setStyle(TableStyle([
      ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#F3F6FA")),
      ("BACKGROUND",(0,1),(-1,1),colors.HexColor("#D9EAF7")),
      ("TEXTCOLOR",(0,0),(-1,-1),colors.HexColor("#17365D")),
      ("FONTNAME",(0,1),(-1,1),"Helvetica-Bold"),("FONTSIZE",(0,1),(-1,1),12),
      ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
      ("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#C9D6E3")),
    ]))
    story.extend([summary_table,Spacer(1,5*mm)])

    def pp(value,style=cell_style):
        return Paragraph(escape(str(value or "")),style)

    table_headers=["OT","Área","Línea","Equipo","Descripción","Actividad","Plan","Crit.","Condición","Pers.","Min","H-H","Origen","Estado"]
    data=[[pp(h,small_bold) for h in table_headers]]
    for row in rows:
        data.append([
          pp(row.get("numero_orden"),cell_center),
          pp(row.get("area_nombre")),
          pp(row.get("linea_nombre")),
          pp(row.get("activo_codigo")),
          pp(row.get("activo_descripcion")),
          pp(row.get("actividad")),
          pp(row.get("plan_trabajo")),
          pp(row.get("criticidad"),cell_center),
          pp(row.get("condicion")),
          pp("" if row.get("personas_usar") is None else f"{float(row['personas_usar']):.0f}",cell_center),
          pp("" if row.get("tiempo_planeado_min") is None else f"{float(row['tiempo_planeado_min']):.0f}",cell_center),
          pp("" if row.get("hh_programadas") is None else f"{float(row['hh_programadas']):.1f}",cell_center),
          pp(row.get("origen"),cell_center),
          pp(row.get("estado"),cell_center),
        ])

    widths=[17,20,18,23,32,34,34,9,22,9,10,10,12,15]
    detail=Table(data,colWidths=[w*mm for w in widths],repeatRows=1)
    style=[
      ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2F75B5")),
      ("TEXTCOLOR",(0,0),(-1,0),colors.white),
      ("ALIGN",(0,0),(-1,0),"CENTER"),("VALIGN",(0,0),(-1,-1),"TOP"),
      ("GRID",(0,0),(-1,-1),0.25,colors.HexColor("#D9E2F3")),
      ("LEFTPADDING",(0,0),(-1,-1),2),("RIGHTPADDING",(0,0),(-1,-1),2),
      ("TOPPADDING",(0,0),(-1,-1),2.5),("BOTTOMPADDING",(0,0),(-1,-1),2.5),
    ]
    for idx in range(1,len(data)):
        if idx%2==0:style.append(("BACKGROUND",(0,idx),(-1,idx),colors.HexColor("#F8FAFC")))
    detail.setStyle(TableStyle(style))
    story.append(detail)

    def add_page_number(canvas,doc_obj):
        canvas.saveState()
        canvas.setFont("Helvetica",6.5)
        canvas.setFillColor(colors.HexColor("#6B7280"))
        canvas.drawRightString(landscape(A4)[0]-8*mm,5*mm,f"Página {doc_obj.page}")
        canvas.restoreState()

    doc.build(story,onFirstPage=add_page_number,onLaterPages=add_page_number)
    out.seek(0)
    filename=f"programacion_{header['especialidad']}_{header['fecha_desde']:%Y%m%d}_{header['fecha_hasta']:%Y%m%d}_v{header['numero_version']}.pdf"
    return out.getvalue(),filename
