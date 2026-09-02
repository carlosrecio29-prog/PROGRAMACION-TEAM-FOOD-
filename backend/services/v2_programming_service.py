from __future__ import annotations

from datetime import date
from html import escape
from io import BytesIO
from typing import Any

from sqlalchemy import text

from backend.database import get_engine

VALID_SPECIALTIES={"MEC","ELE","MET","SER"}


class V2ProgrammingError(ValueError):
    pass


def _validate_week(date_from:date,date_to:date,specialty:str)->str:
    specialty=specialty.upper()
    if specialty not in VALID_SPECIALTIES:
        raise V2ProgrammingError("Especialidad inválida")
    if date_to<date_from:
        raise V2ProgrammingError("La fecha final no puede ser anterior a la inicial")
    if (date_to-date_from).days>6:
        raise V2ProgrammingError("La programación debe abarcar máximo 7 días")
    return specialty


def _capacity(conn,date_from:date,date_to:date,specialty:str)->dict[str,Any]:
    row=conn.execute(text("""
        SELECT
          count(DISTINCT t.id)::int AS tecnicos,
          round(COALESCE(sum(pt.horas_disponibles),0),2) AS hh_disponibles
        FROM programacion.programacion_tecnico pt
        JOIN programacion.tecnico t ON t.id=pt.tecnico_id
        WHERE pt.fecha BETWEEN :date_from AND :date_to
          AND t.especialidad_efectiva=:specialty
    """),{"date_from":date_from,"date_to":date_to,"specialty":specialty}).mappings().one()
    available=float(row["hh_disponibles"] or 0)
    return {
        "technicians":int(row["tecnicos"] or 0),
        "available":round(available,2),
        "target":round(available*.80,2),
        "reserve":round(available*.20,2),
    }


def get_week_programming(*,date_from:date,date_to:date,specialty:str)->dict[str,Any]:
    specialty=_validate_week(date_from,date_to,specialty)
    with get_engine().connect() as conn:
        capacity=_capacity(conn,date_from,date_to,specialty)
        programming=conn.execute(text("""
            SELECT id,estado,hh_disponibles,hh_objetivo,hh_reserva,creado_en,actualizado_en,emitido_en
            FROM programacion.programacion_semanal_v2
            WHERE semana_inicio=:date_from
              AND semana_fin=:date_to
              AND especialidad=:specialty
        """),{"date_from":date_from,"date_to":date_to,"specialty":specialty}).mappings().first()

        programming_id=int(programming["id"]) if programming else None

        rows=[dict(r) for r in conn.execute(text("""
            WITH candidate AS (
              SELECT
                o.id AS orden_mantenimiento_id,
                o.numero_ot,
                o.titulo,
                o.estado,
                o.especialidad,
                a.codigo AS activo_codigo,
                a.descripcion AS activo_descripcion,
                a.area_codigo,
                root.descripcion AS area_nombre,
                p.id AS plan_trabajo_id,
                p.plan_trabajo,
                p.descripcion_grupo,
                p.numero_personas_efectivo,
                p.tiempo_parada_efectivo_min,
                p.requiere_parada,
                COALESCE(o.tiempo_planeado_min,p.tiempo_ejecucion_min) AS tiempo_min,
                round(
                  COALESCE(o.tiempo_planeado_min,p.tiempo_ejecucion_min)
                  / 60.0 * p.numero_personas_efectivo
                ,2) AS hh
              FROM programacion.orden_mantenimiento o
              JOIN programacion.activo a ON a.id=o.activo_id
              LEFT JOIN programacion.activo root ON root.codigo='BA-'||a.area_codigo
              JOIN programacion.plan_trabajo p ON p.id=o.plan_trabajo_id
              WHERE o.periodo=date_trunc('month',CAST(:date_from AS date))::date
                AND o.especialidad=:specialty
                AND upper(COALESCE(o.estado,''))<>'FINALIZADO'
                AND p.numero_personas_efectivo IS NOT NULL
                AND p.tiempo_parada_efectivo_min IS NOT NULL
                AND p.requiere_parada IS NOT NULL
                AND COALESCE(o.tiempo_planeado_min,p.tiempo_ejecucion_min) IS NOT NULL
            )
            SELECT
              c.*,
              CASE WHEN current_item.id IS NOT NULL THEN true ELSE false END AS seleccionado
            FROM candidate c
            LEFT JOIN programacion.programacion_item_v2 current_item
              ON current_item.orden_mantenimiento_id=c.orden_mantenimiento_id
             AND current_item.programacion_id=:programming_id
            WHERE NOT EXISTS (
              SELECT 1
              FROM programacion.programacion_item_v2 other_item
              JOIN programacion.programacion_semanal_v2 other_program
                ON other_program.id=other_item.programacion_id
              WHERE other_item.orden_mantenimiento_id=c.orden_mantenimiento_id
                AND (:programming_id IS NULL OR other_program.id<>:programming_id)
            )
            ORDER BY c.area_codigo,c.numero_ot NULLS LAST,c.activo_codigo,c.plan_trabajo
        """),{
            "date_from":date_from,
            "specialty":specialty,
            "programming_id":programming_id,
        }).mappings()]

        selected_hh=round(sum(float(r["hh"] or 0) for r in rows if r["seleccionado"]),2)

    operating=[r for r in rows if r["requiere_parada"] is False]
    stopped=[r for r in rows if r["requiere_parada"] is True]
    return {
        "date_from":str(date_from),
        "date_to":str(date_to),
        "specialty":specialty,
        "capacity":capacity,
        "programming":dict(programming) if programming else None,
        "selected_hh":selected_hh,
        "selected_ids":[int(r["orden_mantenimiento_id"]) for r in rows if r["seleccionado"]],
        "operating":operating,
        "stopped":stopped,
    }


def save_week_programming(
    *,
    date_from:date,
    date_to:date,
    specialty:str,
    order_ids:list[int],
    created_by:str|None=None,
)->dict[str,Any]:
    specialty=_validate_week(date_from,date_to,specialty)
    unique_ids=list(dict.fromkeys(int(x) for x in order_ids))
    if not unique_ids:
        raise V2ProgrammingError("Debes seleccionar al menos una orden/actividad")

    with get_engine().begin() as conn:
        capacity=_capacity(conn,date_from,date_to,specialty)
        if capacity["target"]<=0:
            raise V2ProgrammingError("Esta especialidad no tiene H-H disponibles para la semana seleccionada")

        current_id=conn.execute(text("""
            SELECT id
            FROM programacion.programacion_semanal_v2
            WHERE semana_inicio=:date_from AND semana_fin=:date_to AND especialidad=:specialty
        """),{"date_from":date_from,"date_to":date_to,"specialty":specialty}).scalar_one_or_none()

        rows=[dict(r) for r in conn.execute(text("""
            SELECT
              o.id AS orden_mantenimiento_id,
              o.numero_ot,
              p.requiere_parada,
              round(
                COALESCE(o.tiempo_planeado_min,p.tiempo_ejecucion_min)
                / 60.0 * p.numero_personas_efectivo
              ,2) AS hh
            FROM programacion.orden_mantenimiento o
            JOIN programacion.plan_trabajo p ON p.id=o.plan_trabajo_id
            WHERE o.id=ANY(:ids)
              AND o.periodo=date_trunc('month',CAST(:date_from AS date))::date
              AND o.especialidad=:specialty
              AND upper(COALESCE(o.estado,''))<>'FINALIZADO'
              AND p.numero_personas_efectivo IS NOT NULL
              AND p.tiempo_parada_efectivo_min IS NOT NULL
              AND p.requiere_parada IS NOT NULL
              AND COALESCE(o.tiempo_planeado_min,p.tiempo_ejecucion_min) IS NOT NULL
        """),{"ids":unique_ids,"date_from":date_from,"specialty":specialty}).mappings()]

        if len(rows)!=len(unique_ids):
            raise V2ProgrammingError("Una o más órdenes ya no están disponibles o tienen datos incompletos")

        conflicts=conn.execute(text("""
            SELECT o.numero_ot,ps.semana_inicio,ps.semana_fin,ps.especialidad
            FROM programacion.programacion_item_v2 pi
            JOIN programacion.programacion_semanal_v2 ps ON ps.id=pi.programacion_id
            JOIN programacion.orden_mantenimiento o ON o.id=pi.orden_mantenimiento_id
            WHERE pi.orden_mantenimiento_id=ANY(:ids)
              AND (:current_id IS NULL OR ps.id<>:current_id)
            LIMIT 10
        """),{"ids":unique_ids,"current_id":current_id}).mappings().all()
        if conflicts:
            first=conflicts[0]
            raise V2ProgrammingError(
                f"La OT {first['numero_ot'] or 'SIN ASIGNAR'} ya está programada "
                f"del {first['semana_inicio']} al {first['semana_fin']}"
            )

        total=round(sum(float(r["hh"] or 0) for r in rows),2)
        if total>capacity["target"]+.001:
            raise V2ProgrammingError(
                f"La selección suma {total:.1f} H-H y supera la meta del 80% ({capacity['target']:.1f} H-H)"
            )

        programming_id=conn.execute(text("""
            INSERT INTO programacion.programacion_semanal_v2(
              semana_inicio,semana_fin,especialidad,hh_disponibles,hh_objetivo,hh_reserva,
              estado,creado_por,actualizado_en
            )
            VALUES(:date_from,:date_to,:specialty,:available,:target,:reserve,'GUARDADA',:created_by,now())
            ON CONFLICT(semana_inicio,semana_fin,especialidad)
            DO UPDATE SET
              hh_disponibles=EXCLUDED.hh_disponibles,
              hh_objetivo=EXCLUDED.hh_objetivo,
              hh_reserva=EXCLUDED.hh_reserva,
              estado='GUARDADA',
              creado_por=COALESCE(EXCLUDED.creado_por,programacion.programacion_semanal_v2.creado_por),
              actualizado_en=now()
            RETURNING id
        """),{
            "date_from":date_from,"date_to":date_to,"specialty":specialty,
            "available":capacity["available"],"target":capacity["target"],"reserve":capacity["reserve"],
            "created_by":created_by,
        }).scalar_one()

        conn.execute(text("""
            DELETE FROM programacion.programacion_item_v2
            WHERE programacion_id=:programming_id
        """),{"programming_id":programming_id})

        for row in rows:
            conn.execute(text("""
                INSERT INTO programacion.programacion_item_v2(
                  programacion_id,orden_mantenimiento_id,hh_programadas,requiere_parada
                )
                VALUES(:programming_id,:order_id,:hh,:requires_stop)
            """),{
                "programming_id":programming_id,
                "order_id":row["orden_mantenimiento_id"],
                "hh":row["hh"],
                "requires_stop":row["requiere_parada"],
            })

    return {
        "ok":True,
        "programming_id":int(programming_id),
        "hh_available":capacity["available"],
        "hh_target":capacity["target"],
        "hh_reserve":capacity["reserve"],
        "hh_programmed":total,
        "progress_pct":round(total/capacity["target"]*100,1) if capacity["target"] else 0,
        "items":len(rows),
    }


def _report_data(programming_id:int)->tuple[dict[str,Any],list[dict[str,Any]]]:
    with get_engine().connect() as conn:
        header=conn.execute(text("""
            SELECT
              ps.id,ps.semana_inicio,ps.semana_fin,ps.especialidad,
              ps.hh_disponibles,ps.hh_objetivo,ps.hh_reserva,ps.estado,
              ps.creado_por,ps.creado_en,ps.actualizado_en,
              round(COALESCE(sum(pi.hh_programadas),0),2) AS hh_programadas,
              count(pi.id)::int AS items
            FROM programacion.programacion_semanal_v2 ps
            LEFT JOIN programacion.programacion_item_v2 pi ON pi.programacion_id=ps.id
            WHERE ps.id=:id
            GROUP BY ps.id
        """),{"id":programming_id}).mappings().first()
        if not header:
            raise V2ProgrammingError("Programación no encontrada")

        rows=[dict(r) for r in conn.execute(text("""
            SELECT
              o.numero_ot,
              a.area_codigo,
              COALESCE(root.descripcion,a.area_codigo) AS area_nombre,
              a.codigo AS activo_codigo,
              a.descripcion AS activo_descripcion,
              p.descripcion_grupo,
              p.plan_trabajo,
              p.numero_personas_efectivo AS personas,
              COALESCE(o.tiempo_planeado_min,p.tiempo_ejecucion_min) AS tiempo_min,
              pi.hh_programadas,
              pi.requiere_parada,
              CASE WHEN pi.requiere_parada THEN 'EQUIPO DETENIDO' ELSE 'EQUIPO FUNCIONANDO' END AS condicion
            FROM programacion.programacion_item_v2 pi
            JOIN programacion.orden_mantenimiento o ON o.id=pi.orden_mantenimiento_id
            JOIN programacion.activo a ON a.id=o.activo_id
            LEFT JOIN programacion.activo root ON root.codigo='BA-'||a.area_codigo
            JOIN programacion.plan_trabajo p ON p.id=o.plan_trabajo_id
            WHERE pi.programacion_id=:id
            ORDER BY pi.requiere_parada,a.area_codigo,o.numero_ot NULLS LAST,a.codigo
        """),{"id":programming_id}).mappings()]

    return dict(header),rows


def export_weekly_excel(programming_id:int)->tuple[bytes,str]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header,rows=_report_data(programming_id)
    wb=Workbook()
    ws=wb.active
    ws.title="Programación semanal"
    ws.sheet_view.showGridLines=False
    ws.freeze_panes="A9"
    ws.page_setup.orientation="landscape"
    ws.page_setup.fitToWidth=1
    ws.sheet_properties.pageSetUpPr.fitToPage=True

    navy="17365D";blue="2F75B5";light="D9EAF7";pale="F3F6FA";white="FFFFFF";dark="1F2937";gray="6B7280"
    thin=Side(style="thin",color="D9E2F3")

    ws.merge_cells("A1:J1")
    ws["A1"]="PROGRAMACIÓN SEMANAL DE MANTENIMIENTO"
    ws["A1"].font=Font(size=18,bold=True,color=white)
    ws["A1"].fill=PatternFill("solid",fgColor=navy)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=30

    ws.merge_cells("A2:J2")
    ws["A2"]=(
      f"Semana {header['semana_inicio']:%d/%m/%Y} al {header['semana_fin']:%d/%m/%Y}  |  "
      f"Especialidad: {header['especialidad']}"
    )
    ws["A2"].font=Font(size=10,bold=True,color=dark)
    ws["A2"].alignment=Alignment(horizontal="center")

    cards=[
      ("A4:B4","A5:B6","H-H DISPONIBLES",float(header["hh_disponibles"] or 0)),
      ("C4:D4","C5:D6","META PROGRAMABLE 80%",float(header["hh_objetivo"] or 0)),
      ("E4:F4","E5:F6","H-H A RESPONDER",float(header["hh_programadas"] or 0)),
      ("G4:H4","G5:H6","RESERVA 20%",float(header["hh_reserva"] or 0)),
      ("I4:J4","I5:J6","ÓRDENES / ACTIVIDADES",int(header["items"] or 0)),
    ]
    for title_range,value_range,label,value in cards:
        ws.merge_cells(title_range);ws.merge_cells(value_range)
        tc=ws[title_range.split(":")[0]];vc=ws[value_range.split(":")[0]]
        tc.value=label;tc.font=Font(size=9,bold=True,color=gray);tc.fill=PatternFill("solid",fgColor=pale);tc.alignment=Alignment(horizontal="center")
        vc.value=value;vc.font=Font(size=16,bold=True,color=navy);vc.fill=PatternFill("solid",fgColor=light);vc.alignment=Alignment(horizontal="center",vertical="center")
        if isinstance(value,float):vc.number_format='0.0'

    headers=["OT","Área","Código equipo","Descripción equipo","Plan de trabajo","Condición","Personas","Tiempo min","H-H","Grupo"]
    for col,label in enumerate(headers,1):
        cell=ws.cell(8,col,label)
        cell.font=Font(bold=True,color=white,size=9)
        cell.fill=PatternFill("solid",fgColor=blue)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=Border(left=thin,right=thin,top=thin,bottom=thin)

    for idx,row in enumerate(rows,9):
        values=[
          row.get("numero_ot") or "SIN ASIGNAR",
          row.get("area_nombre") or row.get("area_codigo") or "",
          row.get("activo_codigo") or "",
          row.get("activo_descripcion") or "",
          row.get("plan_trabajo") or "",
          row.get("condicion") or "",
          float(row["personas"]) if row.get("personas") is not None else "",
          float(row["tiempo_min"]) if row.get("tiempo_min") is not None else "",
          float(row["hh_programadas"]) if row.get("hh_programadas") is not None else "",
          row.get("descripcion_grupo") or "",
        ]
        for col,value in enumerate(values,1):
            cell=ws.cell(idx,col,value)
            cell.font=Font(size=8,color=dark)
            cell.alignment=Alignment(vertical="top",wrap_text=True)
            cell.border=Border(left=thin,right=thin,top=thin,bottom=thin)
            if idx%2==0:cell.fill=PatternFill("solid",fgColor="F8FAFC")
        ws.cell(idx,7).number_format='0'
        ws.cell(idx,8).number_format='0.0'
        ws.cell(idx,9).number_format='0.0'

    widths=[17,24,22,38,38,20,10,12,10,22]
    for i,width in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=width
    ws.auto_filter.ref=f"A8:J{max(8,8+len(rows))}"

    out=BytesIO();wb.save(out);out.seek(0)
    filename=f"programacion_{header['especialidad']}_{header['semana_inicio']:%Y%m%d}_{header['semana_fin']:%Y%m%d}.xlsx"
    return out.getvalue(),filename


def export_weekly_pdf(programming_id:int)->tuple[bytes,str]:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    header,rows=_report_data(programming_id)
    out=BytesIO()
    doc=SimpleDocTemplate(
        out,pagesize=landscape(A4),leftMargin=8*mm,rightMargin=8*mm,
        topMargin=8*mm,bottomMargin=8*mm,title="Programación semanal de mantenimiento"
    )
    styles=getSampleStyleSheet()
    title_style=ParagraphStyle("Title2",parent=styles["Heading1"],fontName="Helvetica-Bold",fontSize=15,leading=18,textColor=colors.HexColor("#17365D"),alignment=TA_CENTER)
    sub_style=ParagraphStyle("Sub2",parent=styles["Normal"],fontSize=8,leading=10,textColor=colors.HexColor("#4B5563"),alignment=TA_CENTER)
    cell=ParagraphStyle("Cell2",parent=styles["Normal"],fontSize=5.8,leading=7)
    cell_center=ParagraphStyle("CellCenter2",parent=cell,alignment=TA_CENTER)
    bold=ParagraphStyle("Bold2",parent=cell,fontName="Helvetica-Bold")

    story=[
      Paragraph("PROGRAMACIÓN SEMANAL DE MANTENIMIENTO",title_style),
      Paragraph(
        f"Semana {header['semana_inicio']:%d/%m/%Y} al {header['semana_fin']:%d/%m/%Y} &nbsp;|&nbsp; "
        f"Especialidad: <b>{header['especialidad']}</b>",
        sub_style
      ),
      Spacer(1,3*mm)
    ]

    summary=[
      [Paragraph("H-H DISPONIBLES",bold),Paragraph("META 80%",bold),Paragraph("H-H A RESPONDER",bold),Paragraph("RESERVA 20%",bold),Paragraph("ACTIVIDADES",bold)],
      [f"{float(header['hh_disponibles'] or 0):.1f}",f"{float(header['hh_objetivo'] or 0):.1f}",f"{float(header['hh_programadas'] or 0):.1f}",f"{float(header['hh_reserva'] or 0):.1f}",str(int(header['items'] or 0))]
    ]
    st=Table(summary,colWidths=[52*mm,52*mm,52*mm,52*mm,52*mm],rowHeights=[7*mm,8*mm])
    st.setStyle(TableStyle([
      ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#F3F6FA")),
      ("BACKGROUND",(0,1),(-1,1),colors.HexColor("#D9EAF7")),
      ("TEXTCOLOR",(0,0),(-1,-1),colors.HexColor("#17365D")),
      ("FONTNAME",(0,1),(-1,1),"Helvetica-Bold"),("FONTSIZE",(0,1),(-1,1),12),
      ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
      ("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#C9D6E3")),
    ]))
    story.extend([st,Spacer(1,4*mm)])

    def pp(v,style=cell):
        return Paragraph(escape(str(v or "")),style)

    data=[[pp(x,bold) for x in ["OT","Área","Equipo","Descripción","Plan de trabajo","Condición","Pers.","Min","H-H"]]]
    for row in rows:
        data.append([
          pp(row.get("numero_ot") or "SIN ASIGNAR",cell_center),
          pp(row.get("area_nombre") or row.get("area_codigo")),
          pp(row.get("activo_codigo")),
          pp(row.get("activo_descripcion")),
          pp(row.get("plan_trabajo")),
          pp(row.get("condicion")),
          pp("" if row.get("personas") is None else f"{float(row['personas']):.0f}",cell_center),
          pp("" if row.get("tiempo_min") is None else f"{float(row['tiempo_min']):.0f}",cell_center),
          pp("" if row.get("hh_programadas") is None else f"{float(row['hh_programadas']):.1f}",cell_center),
        ])
    table=Table(data,colWidths=[24*mm,28*mm,27*mm,44*mm,55*mm,30*mm,10*mm,12*mm,12*mm],repeatRows=1)
    table.setStyle(TableStyle([
      ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2F75B5")),
      ("TEXTCOLOR",(0,0),(-1,0),colors.white),
      ("VALIGN",(0,0),(-1,-1),"TOP"),
      ("GRID",(0,0),(-1,-1),0.25,colors.HexColor("#D9E2F3")),
      ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F8FAFC")]),
      ("LEFTPADDING",(0,0),(-1,-1),2.5),("RIGHTPADDING",(0,0),(-1,-1),2.5),
      ("TOPPADDING",(0,0),(-1,-1),2.5),("BOTTOMPADDING",(0,0),(-1,-1),2.5),
    ]))
    story.append(table)
    doc.build(story)
    out.seek(0)
    filename=f"programacion_{header['especialidad']}_{header['semana_inicio']:%Y%m%d}_{header['semana_fin']:%Y%m%d}.pdf"
    return out.getvalue(),filename
