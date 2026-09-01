from io import BytesIO
from openpyxl import Workbook
from backend.parsers.operational import parse_order_states, parse_monthly_planning
from backend.parsers.team_food import parse_team_food

def xlsx(rows):
    wb=Workbook(); ws=wb.active
    for row in rows: ws.append(row)
    b=BytesIO(); wb.save(b); return b.getvalue()

def test_order_states():
    p=parse_order_states(xlsx([["Orden","Estado"],[123,"FINALIZADA"],[124,"PENDIENTE"]]))
    assert len(p.rows)==2 and p.rows[0]["order_number"]=="123"

def test_planning():
    headers=["Activo","Especialidad","Orden","PlanTrabajo","TiempoPlaneado","Estado","FechaPlaneadaInicio"]
    p=parse_monthly_planning(xlsx([headers,["A1","MEC","SIN ASIGNAR","100-RUTINA",60,"PENDIENTE","2026-09-01"]]))
    assert len(p.rows)==1 and p.rows[0]["plan_canonical"]=="RUTINA"


def test_team_food_preserves_unknown_plan_data():
    wb=Workbook()
    assets=wb.active;assets.title="EQUIPOS PLANTA BARRANQUILLA"
    assets.append(["Código","Descripción","ActivoPadre","Criticidad","Especialidad","Departamento","Agrupacion","Estado"])
    assets.append(["A1","BOMBA 1","", "A","MEC","JABONERÍA","LINEA 1","Habilitado."])
    planning=wb.create_sheet("PLANEACION")
    planning.append(["Activo","Descripción","IdCronogramaPlaneacion","PlanTrabajo","Estado"])
    planning.append(["A1","RUTINA MENSUAL","1","10-RUTINA MENSUAL","Habilitado."])
    plans=wb.create_sheet("PLAN DE TRABAJO")
    plans.append(["Grupo","DescripcionGrupo","PlanTrabajo","TiempoEjecucion","NumeroPersonas","TiempoParada","Especialidad","Estado"])
    plans.append([10,"BOMBAS","RUTINA MENSUAL",60,None,None,"MEC","Habilitado."])
    orders=wb.create_sheet("ORDENES MENSUALES")
    orders.append(["MES","ÁREA","# DE ORDEN","ESPECIALIDAD","CÓDIGO","DESCRIPCIÓN","OBSERVACIÓN","ESTADO","FECHA PROG","CRITICIDAD","AÑO","TIEMPO"])
    orders.append(["SEPTIEMBRE","JABONERÍA","OT-1-26","MEC","A1","BOMBA 1","RUTINA MENSUAL","PENDIENTE","2026-09-01","A",2026,60])
    technicians=wb.create_sheet("ESPECIALIDAD DE CADA TECNICO ")
    technicians.append(["NOMBRE","Especialiidad "])
    technicians.append(["MARIO ACOSTA","MECANICO "])
    technicians.append(["MIGUEL OSORIO","REFRI GERACION "])
    b=BytesIO();wb.save(b)
    p=parse_team_food(b.getvalue())
    assert p["plans"].rows[0]["persons"] is None
    assert p["plans"].rows[0]["condition"]=="SIN CLASIFICAR"
    assert p["plans"].rows[0]["plan_key"]=="10-RUTINA MENSUAL"
    assert p["orders"].metadata["program_year"]==2026
    assert [r["specialty"] for r in p["technicians"].rows]==["MEC","SER"]
    assert p["technicians"].rows[1]["name_normalized"]=="MIGUEL OSORIO"


def test_team_food_deduplicates_only_repeated_ot():
    wb=Workbook()
    ws=wb.active;ws.title="ORDENES MENSUALES"
    ws.append(["MES","ÁREA","# DE ORDEN","ESPECIALIDAD","CÓDIGO","DESCRIPCIÓN","OBSERVACIÓN","ESTADO","FECHA PROG","CRITICIDAD","AÑO","TIEMPO"])
    ws.append(["SEPTIEMBRE","A","OT-1","MEC","EQ-1","Equipo 1","RUTINA","PENDIENTE","2026-09-01","A",2026,18])
    ws.append(["SEPTIEMBRE","A","OT-1","MEC","EQ-1","Equipo 1","RUTINA","PENDIENTE","2026-09-01","A",2026,29])
    ws.append(["SEPTIEMBRE","A","OT-2","MEC","EQ-1","Equipo 1","RUTINA","PENDIENTE","2026-09-01","A",2026,18])
    b=BytesIO();wb.save(b)
    parsed=parse_team_food(b.getvalue())
    orders=parsed["orders"].rows
    stats=parsed["orders"].metadata["monthly_summary"]["9"]["MEC"]
    assert len(orders)==2
    assert {r["order_number"] for r in orders}=={"OT-1","OT-2"}
    assert stats["master_rows"]==3
    assert stats["unique_ot"]==2
    assert stats["repeated_extra_rows"]==1
    assert stats["pending_unique_ot"]==2
